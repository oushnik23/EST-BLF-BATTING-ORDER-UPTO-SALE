import pandas as pd
import datetime
import numpy as np
import mysql.connector as msql
from mysql.connector import Error
from google.cloud import bigquery

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

import os 
os.chdir(r"D:\Oushnik Sarkar\Python")
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'D:/Oushnik Sarkar/data-warehousing-prod.json'

#Imports google cloud client library and initiates BQ service
from google.cloud import bigquery
#from google.cloud import datastore
bigquery_client = bigquery.Client()

QUERY = """
SELECT
    Centre, FinYear, Season, SaleNo, AuctionDate, 
    CASE WHEN SaleNo BETWEEN 1 AND 13 THEN 52 + SaleNo ELSE SaleNo END AS SaleAlies,
    LotNo, 
    Garden, GardenMDM, Grade, GradeMDM, InvoiceNo, 
    Buyer, BuyerMDM, BuyerGroup, BrokerCode,
    Seller, SellerGroup, Category, SubCategory, TeaType, 
    SubTeaType, LotStatus, Area, EstBlf, GPDATE, ReprintNo,
    
    SUM(TotalWeight) AS Sold_Qty,
    SUM(Value) AS Total_Value
FROM `data-warehousing-prod.EasyReports.SaleTransactionView`
WHERE
    Season IN (2024,2023) AND Area IN ("AS", "DO", "TR", "CA", "TP") AND FinYear IN ("2023-24","2024-25")
    AND if(SaleNo >=1 and SaleNo <=13, 52+SaleNo,SaleNo) between 14 and 18

GROUP BY Centre, FinYear, Season, SaleNo, AuctionDate, LotNo, Garden, GardenMDM, Grade, GradeMDM, InvoiceNo, Buyer, BuyerMDM, BuyerGroup, 
BrokerCode, Seller, SellerGroup, Category, SubCategory, TeaType, SubTeaType, LotStatus, Area, EstBlf, GPDATE, ReprintNo, SaleAlies"""

Query_Results = bigquery_client.query(QUERY)
df = Query_Results.to_dataframe()

df['Avg_Price'] = df['Total_Value'] / df['Sold_Qty']


df_as_est=df[df['Area'].isin(["AS"]) & df['EstBlf'].isin(["EST"]) & df['Category'].isin(["CTC"])]

df_as_est['Season'].unique()

summary_df = (df_as_est.groupby(["FinYear", "GardenMDM"]).agg({"Sold_Qty":"sum","Avg_Price":"mean"}).reset_index())
unstacked_df = summary_df.set_index(["GardenMDM", "FinYear"]).unstack("FinYear")
pivot_df=unstacked_df.swaplevel(axis=1).sort_index(axis=1)

#-----------------------Create Rank-----------------------#

valid_rows = (pivot_df.xs("Sold_Qty", axis=1, level=1) >= 1000).any(axis=1)

def rank_avg_price(series):
    return series.rank(ascending=False, method="dense")  # Higher Avg_Price = Better Rank

rank_df = pivot_df.xs("Avg_Price", axis=1, level=1)[valid_rows].apply(rank_avg_price)

# Step 3: Insert Rank into `final_df` under each year
for year in rank_df.columns:
    pivot_df[(year, "Rank")] = rank_df[year]

# Step 4: Ensure proper column order
pivot_df = pivot_df.sort_index(axis=1)

desired_order = ['Sold_Qty', 'Avg_Price', 'Rank']

# Rearrange columns under each Garden name
new_columns = []
for garden in pivot_df.columns.levels[0]:  # Iterate over the garden names
    for metric in desired_order:  # Maintain the desired metric order
        new_columns.append((garden, metric))

# Update the DataFrame with the new column order
pivot_df = pivot_df[new_columns]

new_column_order = sorted(pivot_df.columns, key=lambda x: x[0], reverse=True)  # Sort FinYear in descending order

# Step 2: Reorder the columns in final_df
pivot_df = pivot_df[new_column_order]

#-----------------------Create Grand Total-----------------------#

grand_total = summary_df.groupby("FinYear").agg({"Sold_Qty": "sum", 
    "Avg_Price": lambda x: (x * summary_df.loc[x.index, "Sold_Qty"]).sum() / summary_df.loc[x.index, "Sold_Qty"].sum()})

# Reshape Grand Total to match the pivot table
grand_total = grand_total.unstack().to_frame().T  # Convert to a row format
grand_total.index = ["Grand Total"]  # Rename index

grand_total=grand_total.swaplevel(axis=1).sort_index(axis=1)

# Append Grand Total as a new row
final_df = (pd.concat([pivot_df, grand_total], axis=0)).round(2)

#----------------Sorting Based on Rank----------------#
grand_total_row = final_df.loc["Grand Total"]

'''
sorted_df = final_df.drop(index="Grand Total").sort_values(by=("2024-25", "Rank"), ascending=True)
sorted_df = pd.concat([sorted_df, grand_total_row.to_frame().T])
'''
#----------------Sorting Based on Rank and Avg Price----------------#

# Separate rows with and without a rank in 2024-25
has_rank = final_df.drop(index="Grand Total").loc[final_df[("2024-25", "Rank")].notna()]
no_rank = final_df.drop(index="Grand Total").loc[final_df[("2024-25", "Rank")].isna()]

# Sort rows with a rank by 2024-25 Rank in ascending order
has_rank_sorted = has_rank.sort_values(by=("2024-25", "Rank"), ascending=True)

# Sort rows without a rank by 2024-25 Avg Price in descending order
no_rank_sorted = no_rank.sort_values(by=("2024-25", "Avg_Price"), ascending=False)

# Combine the two sorted DataFrames and append the Grand Total row at the end
sorted_df = pd.concat([has_rank_sorted, no_rank_sorted, grand_total_row.to_frame().T])

#-----------------------Create Calculation-----------------------#

sorted_df[("Diff", "Sold_Qty")] = (
    sorted_df[("2024-25", "Sold_Qty")] - sorted_df[("2023-24", "Sold_Qty")])

sorted_df[("Diff", "Avg_Price")] = (
    sorted_df[("2024-25", "Avg_Price")] - sorted_df[("2023-24", "Avg_Price")])

#✅ Rename columns to match report style
sorted_df.columns = pd.MultiIndex.from_tuples([
    ("2024-25", "Sold Qty"),("2024-25", "Avg Price"),("2024-25", "Rank"),
    
    ("2023-24", "Sold Qty"),("2023-24", "Avg Price"),("2023-24", "Rank"),
    
    ("Difference", "Diff Qty"),("Difference", "Diff Price"),
])

def format_number(x):
    if pd.isna(x):
        return ""
    return f"{float(x):,}"

# Format selected columns
for col in sorted_df.columns:
    if "Qty" in col[1] or "Price" in col[1]:
        sorted_df[col] = sorted_df[col].apply(format_number)
    elif "Rank" in col[1]:
        sorted_df[col] = sorted_df[col].astype("Int64") 
        
######## Important ########

sorted_df.columns = sorted_df.columns.set_names(["Season", "Garden"])

# Remove commas and convert 'Sold Qty' columns to numeric
sold_qty_2024 = sorted_df[("2024-25", "Sold Qty")].replace("", "0").str.replace(",", "").astype(float)
sold_qty_2023 = sorted_df[("2023-24", "Sold Qty")].replace("", "0").str.replace(",", "").astype(float)

# Create a mask for rows where Sold Qty is 0 for both years
mask = (sold_qty_2024 == 0) & (sold_qty_2023 == 0)

# Filter out rows matching the condition while keeping MultiIndex intact
filtered_df = sorted_df[~mask]

#print(filtered_df.columns)
#print(filtered_df.index)
#print(sorted_df.columns.levels)


#-----------------------------------------DESIGN-----------------------------------------#

excel_path = "final_output.xlsx"
filtered_df.to_excel(excel_path, index=True,sheet_name="AS EST")

# Now open the Excel file and delete the 3rd row (index starts from 1 in openpyxl)
wb = load_workbook(excel_path)
ws = wb.active

# Number of header rows from MultiIndex
header_rows = 2

thin_border = Border(left=Side(style='thin'),
    right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=9):  # A=1, I=9
    for cell in row:
        cell.border = thin_border
        cell.font=Font(size=11)

#-----------------------------------------Make color-------------------------------------#

# Number of header rows from MultiIndex
header_rows = 2

header_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
index_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

# Color the MultiIndex header rows (should be rows 1 and 2 at this point)
for row in ws.iter_rows(min_row=1, max_row=header_rows, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format="#,##,##0"

# Color the index column (A), from data start row to bottom
data_start_row = header_rows + 1

for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.fill = index_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
#------------------------------------------------------------------------------#

for cell in ws['A']:
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.font=Font(bold=True)
     
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length

    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width
    
for row in ws['B:I']:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
ws.insert_rows(1)

# Set the heading in cell A1
heading_cell = ws.cell(row=1, column=1)  # Cell A1
heading_cell.value = "Batting Order"
heading_cell.font = Font(size=12, bold=True)
heading_cell.alignment = Alignment(horizontal='left', vertical='center')

# Merge cells across A1 to I1
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

# Adjust row height for the heading
ws.row_dimensions[1].height = 20

ws.insert_rows(2, amount=9)  # Insert rows starting at row 2


Area=','.join(df_as_est['Area'].unique())
Estblf=','.join(df_as_est['EstBlf'].unique())
centre = ", ".join(df_as_est["Centre"].unique())
df_as_est['FinYear'].max()
df_as_est['FinYear'].min()
df_as_est['Season'].max()
df_as_est['Season'].min()
df_as_est['Category'].unique()


value = df_as_est['SaleAlies'].max() - 52 if df_as_est['SaleAlies'].max() > 52 else df_as_est['SaleAlies'].max()


# Content to add
info_text = [
    f"Area: {Area}",
    f"Est / Blf: {Estblf}",
    f"Centre: {centre}",
    f"Current Season: Season- {df_as_est['Season'].max()} FY- {df_as_est['FinYear'].max()}",
    f"Previous Season: Season- {df_as_est['Season'].min()} FY- {df_as_est['FinYear'].min()}",
    "From Sale No: 14",
    f"Category: {','.join(df_as_est['Category'].unique())}",
    f"To Sale No: {value}",
    "Cut Off Qty: 1000",]

# Write each line into column A, spanning A2:A10
for idx, line in enumerate(info_text, start=2):
    cell = ws.cell(row=idx, column=1)  # Write in column A
    cell.value = line
    cell.font = Font(size=12)  # Set font size
    cell.alignment = Alignment(horizontal='left', vertical='center')

# Merge cells across A to I for each line
for row in range(2, 2 + len(info_text)):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

# Apply number formatting to remove decimal places but retain full precision
            
import re

for col in ws.iter_cols(min_col=8, max_col=9, min_row=data_start_row + 11, max_row=ws.max_row):  # Columns H and I
    for cell in col:
        try:
            if isinstance(cell.value, str):
                # Clean up the cell value by removing commas, spaces, and non-numeric characters
                cleaned = cell.value.replace(',', '').replace(' ', '')
                cleaned = re.sub(r'[^0-9\.]', '', cleaned)  # Remove any non-numeric characters

                if cleaned.replace('.', '', 1).isdigit():  # Check if cleaned string is numeric
                    cell.value = float(cleaned)  # Convert to float if valid

            # If it's a float or int, keep the decimal value internally
            if isinstance(cell.value, (int, float)):
                # Hide decimals in the output using a custom number format
                cell.number_format = '0'  # Format without decimals for display purposes

                # Internally retain decimal value for future reference
                if isinstance(cell.value, float):
                    # Keep the original value (store as float internally)
                    pass  # No action needed, value remains as float

        except Exception as e:
            print(f"Skipping cell {cell.coordinate}: {e}")
            
for col in ws.iter_cols(min_col=2, max_col=6, min_row=data_start_row + 11, max_row=ws.max_row):
    for cell in col:
        try:
            if isinstance(cell.value, str):
                cleaned = cell.value.replace(',', '').replace(' ', '')
                cleaned = re.sub(r'[^0-9\.]', '', cleaned)
                if cleaned.replace('.', '', 1).isdigit():
                    cell.value = float(cleaned)  # Keep as float

            if isinstance(cell.value, (int, float)):
                # Visually show only the integer part, keep float value
                cell.number_format = '#,##,##0'  # 👈 This hides decimals in display

        except Exception as e:
            print(f"Skipping cell {cell.coordinate}: {e}")
            
# Adjust row heights for better visibility
for row in range(2, 2 + len(info_text)):
    ws.row_dimensions[row].height = 15

for cell in ws[ws.max_row]:
    cell.font = Font(bold=True)
    
len(info_text)
#ws.merge_cells('A1:I1')

ws.unmerge_cells('A1:I1')
ws.merge_cells('B11:D11')
ws.merge_cells('E11:G11')
ws.merge_cells('H11:I11')

ws.delete_rows(13)

ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.paperSize = ws.PAPERSIZE_A4

# Fit the entire width on one page, height can flow to multiple
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0  # Unlimited pages vertically

# Center horizontally on page
ws.page_setup.horizontalCentered = True

# Enable fit-to-page setting
from openpyxl.worksheet.properties import PageSetupProperties
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

# Set tight margins for better use of space
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4
ws.page_margins.header = 0.2
ws.page_margins.footer = 0.2

# Optional: print with gridlines
ws.print_options.gridLines = True

wb.save(excel_path)

print("✅ Excel file created successfully!")

#-----------------------------------SENDING EMAIL-----------------------------------#

import smtplib
from email.message import EmailMessage

# Email details
sender_email = "website@parcon.in"
receiver_email = "mis@parcon-india.com"
subject = "AS EST File"
body = "Hi, please find the attached Excel file."

# Create email
msg = EmailMessage()
msg['Subject'] = subject
msg['From'] = sender_email
msg['To'] = receiver_email
msg.set_content(body)

# Attach the Excel file
with open('final_output.xlsx', 'rb') as f:
    file_data = f.read()
    file_name = 'final_output.xlsx'

msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=file_name)

# SMTP server setup (Example for Gmail)
smtp_server = "smtp.gmail.com"
smtp_port = 587

# Send email
with smtplib.SMTP(smtp_server, smtp_port) as server:
    server.starttls()
    server.login(sender_email, "plje epzu hmvc eugz")  # Replace with your real password
    server.send_message(msg)

print("✅ Email sent successfully!")